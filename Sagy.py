import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import json

st.title("AKI - Make your Reports easier :scroll:")
uploaded_file = st.file_uploader("Choose a TXT file :scroll:", type=["txt"])

if uploaded_file:

    with open("Brands.json", "r") as json_file:
        brands_data = json.load(json_file)

    data = uploaded_file.read().decode("utf-8").splitlines()
    
    df = pd.DataFrame(data, columns=["Column1"])

    st.subheader("Original Data")
    st.write(df.head())

    df = df.iloc[20:]
    split_data = df['Column1'].str.split('$', expand=True)

    total_rows = split_data[split_data.apply(lambda row: row.astype(str).str.contains(r'\s*Total', case=False, regex=True).any(), axis=1)].index
    rows_to_remove = total_rows.union(total_rows + 1)
    split_data = split_data.drop(rows_to_remove)

    split_data = split_data.drop(columns=[5, 6, 7], errors='ignore')


    split_data[0] = split_data[0].str.lstrip()

    split_data[5] = split_data[0].str[:3]


    for col in [1, 2, 3, 4]: 
        split_data[col] = split_data[col].str.strip()  
        split_data[col] = split_data[col].str.replace(r'\s{2,}', ' ', regex=True)


    split_data[6] = split_data[5].apply(lambda x: brands_data.get(x, '') if x in brands_data else '')

    st.subheader("Processed Data")
    st.write(split_data)

    # إضافة أسماء الأعمدة الجديدة
    split_data.columns = ['Item', 'Des', 'QTY', 'BNS', 'Total', 'Brands', 'SC']

    # طباعة البيانات لمعرفة الفهارس
    st.write("Data before deleting the specified row:")
    st.write(split_data)

    # حذف السطر الثالث بناءً على القيم المحددة
    row_to_remove = split_data[(split_data['Item'] == 'Item') & 
                                (split_data['Des'] == 'Desc') & 
                                (split_data['QTY'] == 'Qty') & 
                                (split_data['BNS'] == 'BNS') & 
                                (split_data['Total'] == 'Amount') & 
                                (split_data['Brands'] == 'Ite')]

    if not row_to_remove.empty:
        split_data = split_data.drop(row_to_remove.index)

    # إضافة سطر فارغ لتجنب تطبيق الدوال عليه
    empty_row = pd.Series([''] * len(split_data.columns), index=split_data.columns)
    split_data = pd.concat([pd.DataFrame([empty_row]), split_data], ignore_index=True)

    output_file = "processed_data.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    # إعداد تنسيق الخط واللون للسطر الأول
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    header_font = Font(bold=True)

    # تعيين القيم الثابتة في السطر الأول
    for idx, col_name in enumerate(split_data.columns, start=1):
        sheet.cell(row=1, column=idx, value=col_name).font = header_font
        sheet.cell(row=1, column=idx).fill = header_fill

    # إضافة البيانات إلى الملف
    for row in range(split_data.shape[0]):
        for col in range(split_data.shape[1]):
            cell = sheet.cell(row=row + 2, column=col + 1, value=split_data.iat[row, col])
            # إضافة الحدود لكل الخلايا
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # ضبط عرض الأعمدة لتناسب المحتوى
    for col in range(split_data.shape[1]):
        max_length = 0
        for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col + 1, max_col=col + 1):
            for item in cell:
                max_length = max(max_length, len(str(item.value)))

        adjusted_width = (max_length + 2)  # إضافة بعض المساحة الإضافية
        sheet.column_dimensions[sheet.cell(row=2, column=col + 1).column_letter].width = adjusted_width

    workbook.save(output_file)

    st.download_button(
        label="Download Processed Excel File",
        data=open(output_file, "rb").read(),
        file_name=output_file,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
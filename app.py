import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import tempfile
import json


def create_new_excel(uploaded_file):
    xls = pd.ExcelFile(uploaded_file)
    
    workbook = Workbook()
    sheet = workbook.active

    with open('Brands.json', 'r', encoding='utf-8') as json_file:
        brands_data = json.load(json_file)

    header = [
        "Org", "Item Code", "Brand", "Exp resp", "UTL(Y/N)", "SC Planner",
        "Item Descr", "Status", "CLASS 5", 
        "0-3 Mths", "4-6 Mths", "7-12 Mths", "> 12 Mths", 
        "Ageing", "Total", 
        "0-3Mth", "4-6Mth", "7-12Mth", ">12Mth", "Total",
        "Avg", "AVGs", "Mths", "Units", "Near exp/V"
    ]

    sheet.append([""] * len(header)) 
    sheet.append([""] * len(header))  
    sheet.append(header)                
    sheet.freeze_panes = 'A4'
    for cell in sheet[3]:  
        cell.font = Font(bold=True) 
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") 

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.border = border

    if "DPH,APH" in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name="DPH,APH", skiprows=6)  
 
        st.write(f"Columns in sheet DPH,APH: {df.columns.tolist()}")
        
  
        df.columns = df.columns.str.strip()
        
        df = df[~df['Item Descr'].isin(['Other', 'Tota'])]

        for index, row in df.iterrows():
            try:
                brand_name = row[1]
                person_name = brands_data.get(brand_name, "")  

                new_row = [
                    "PVT",                  
                    row[0],                
                    brand_name,           
                    "",                    
                    "",                   
                    person_name,       
                    row[2],              
                    "",                    
                    "",                  
                    row[3],                
                    row[4],              
                    row[5],               
                    row[6],             
                    "",                    
                    row[7],              
                    row[8],            
                    row[9],               
                    row[10],              
                    row[11],              
                    row[12],                    
                    row[13],             
                    "",                     
                    row[14],                
                    "",                     
                    "",                     
                ]
                sheet.append(new_row)

                row_index1 = sheet.max_row 
                sheet[f'N{row_index1}'] = f'=SUM(L{row_index1}:M{row_index1})'
                sheet[f'V{row_index1}'] = f'=(O{row_index1}/T{row_index1})*U{row_index1}' 
                sheet[f'Y{row_index1}'] = f'=(O{row_index1}/T{row_index1})*X{row_index1}'  

            except Exception as e:
                st.error(f"Error while processing row {index}: {str(e)}")

    if "DPI,API" in xls.sheet_names:
        df_api = pd.read_excel(xls, sheet_name="DPI,API", skiprows=6)  

        df_api.columns = df_api.columns.str.strip()

        df_api = df_api[~df_api['Item Descr'].isin(['Other', 'Tota'])]

        for index, row in df_api.iterrows():
            try:
                brand_name = row[1]  
                person_name = brands_data.get(brand_name, "")

                new_row_api = [
                    "INST",                 
                    row[0],               
                    brand_name,             
                    "",                  
                    "",                 
                    person_name,                 
                    row[2],            
                    "",                   
                    "",                  
                    row[3],              
                    row[4],              
                    row[5],             
                    row[6],             
                    "",               
                    row[7],               
                    row[8],              
                    row[9],             
                    row[10],             
                    row[11],        
                    row[12],                   
                    row[13],             
                    "",                   
                    row[14],                  
                    "",                    
                    "",                     
                ]
                sheet.append(new_row_api)

                row_index = sheet.max_row 
                sheet[f'N{row_index}'] = f'=SUM(L{row_index}:M{row_index})'
                sheet[f'V{row_index}'] = f'=(O{row_index}/T{row_index})*U{row_index}' 
                sheet[f'Y{row_index}'] = f'=(O{row_index}/T{row_index})*X{row_index}' 

                sheet['N1'] = "Ageing"
                sheet['O1'] = "Total"
                sheet['P1'] = "Coverage"
                sheet['Q1'] = "Ageing %"
                sheet['R1'] = "12M"
                sheet['V1'] = "Average Sales"
                sheet['M2'] = "=SUBTOTAL(9,M3:M" + str(sheet.max_row) + ")"
                sheet['N2'] = "=SUBTOTAL(9,N3:N" + str(sheet.max_row) + ")"
                sheet['O2'] = "=SUBTOTAL(9,O3:O" + str(sheet.max_row) + ")"
                sheet['P2'] = "=O2/V2"
                sheet['Q2'] = "=N2/O2"
                sheet['R2'] = "=M2/O2"
                sheet['T2'] = "=SUBTOTAL(9,T3:T" + str(sheet.max_row) + ")"
                sheet['U2'] = "=SUBTOTAL(9,U3:U" + str(sheet.max_row) + ")"
                sheet['V2'] = "=SUBTOTAL(9,V3:V" + str(sheet.max_row) + ")"
                sheet['W2'] = "=T2/U2"
                sheet['Y2'] = "=SUBTOTAL(9,Y3:Y" + str(sheet.max_row) + ")"

                for cell in ['V2', 'M2', 'N2', 'O2', 'P2', 'Q2', 'R2', "S2", 'T2', 'U2', 'W2', 'X2', 'Y2']:
                    sheet[cell].border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for cell in ['V1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', "S1", 'T1', 'U1', 'W1', 'X1', 'Y1']:
                    sheet[cell].border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for cell in sheet[1]: 
                     cell.alignment = Alignment(horizontal='center')
                for cell in sheet[2]: 
                     cell.alignment = Alignment(horizontal='center')

            except Exception as e:
                st.error(f"Error while processing row {index}: {str(e)}")

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        last_row = sheet.max_row
        last_column = sheet.max_column
        for row in sheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=last_column):
            for cell in row:
                cell.border = border
    sheet.auto_filter.ref = f"A3:{chr(65 + len(header) - 1)}3" 

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    workbook.save(temp_file.name)

    return temp_file.name

uploaded_file = st.file_uploader("Choose a Stock Ageing Report (Excel file)", type=["xlsx"])

if uploaded_file is not None:
    output_path = create_new_excel(uploaded_file)

    st.success("New Excel file created successfully!")
    with open(output_path, "rb") as f:
        st.download_button(
            label="Download Edited Excel File",
            data=f,
            file_name="Stock_Ageing_Report_v01.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

